import urllib.request
import xml.etree.ElementTree as ET
import openpyxl as excel

# Web APIからXMLデータを取得 --- (*1)
api = 'https://api.aoikujira.com/tenki/week.php?fmt=xml&city=319'
req = urllib.request.Request(api)
with urllib.request.urlopen(req) as res:
    xml_str = res.read().decode("utf-8")

# 文字列からXMLデータを読み込む --- (*2)
xml = ET.fromstring(xml_str)

# Excelブックに書き込む --- (*3)
# 新規ブックを作りシートを得る
book = excel.Workbook()
sheet = book.active

# XMLデータを順に得る --- (*4)
i = 1
for c in xml.iter('date'):
    date = c.attrib['value'] # --- (*5)
    forecast = c[0].text # --- (*6)
    sheet.cell(row=i, column=1).value = date
    sheet.cell(row=i, column=2).value = forecast
    i += 1

# ファイルを保存 --- (*7)
book.save("webapi-xml2.xlsx")


