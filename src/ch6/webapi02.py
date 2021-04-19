import urllib.request, json
import openpyxl as excel

# Web APIからJSONデータを取得 --- (*1)
api = 'https://api.aoikujira.com/tenki/week.php?fmt=json&city=319'
req = urllib.request.Request(api)
with urllib.request.urlopen(req) as res:
    body = json.load(res)

# JSONデータをExcelブックに書き込む --- (*2)
# 新規ブックを作りシートを得る
book = excel.Workbook()
sheet = book.active

# JSONデータを順に得る --- (*3)
for i, row in enumerate(body['319']):
    sheet.cell(row=i+1, column=1).value = row['date']
    sheet.cell(row=i+1, column=2).value = row['forecast']

# ファイルを保存 --- (*4)
book.save("webapi02.xlsx")
