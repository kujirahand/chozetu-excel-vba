import urllib.request
import configparser
import openpyxl as excel

# Web APIからINIデータを取得 --- (*1)
api = 'https://api.aoikujira.com/tenki/week.php?fmt=ini&city=319'
req = urllib.request.Request(api)
with urllib.request.urlopen(req) as res:
    ini_str = res.read().decode("utf-8")
    print(ini_str)

# 文字列からINIデータを読み込む --- (*2)
ini = configparser.ConfigParser(interpolation=None)
ini.read_string(ini_str)

# Excelブックに書き込む --- (*3)
# 新規ブックを作りシートを得る
book = excel.Workbook()
sheet = book.active

# INIデータを順に書き込む --- (*4)
i = 1
sec = ini['319'] # 319のセクションを得る --- (*5)
for key in sec.keys(): # 繰り返す --- (*6)
    date = key + '日'
    forecast = sec.get(key, '') # --- (*7)
    sheet.cell(row=i, column=1).value = date
    sheet.cell(row=i, column=2).value = forecast
    i += 1

# ファイルを保存 --- (*8)
book.save("webapi-ini.xlsx")


